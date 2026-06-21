from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, BackgroundTasks
from fastapi.concurrency import run_in_threadpool
from app.core.database import get_database
from app.core.security import get_current_active_user, serialize_doc, require_permission
from app.core.config import settings
from app.models.product import ProductCreate, ProductUpdate
from bson import ObjectId
from datetime import datetime, timedelta
from typing import Optional
import io
import re
import asyncio
import logging
import base64
import json
import httpx

logger = logging.getLogger(__name__)

router = APIRouter()
ocr_lock = asyncio.Lock()

def parse_expiry_string(expiry_str: Optional[str]) -> Optional[datetime]:
    if not expiry_str:
        return None
    expiry_str = expiry_str.strip()
    if not expiry_str:
        return None
    
    # Try ISO parse first
    try:
        return datetime.fromisoformat(expiry_str.replace("Z", "+00:00"))
    except ValueError:
        pass
        
    # Try MM/YY or MM-YY
    m = re.match(r"^(\d{1,2})[/\-](\d{2})$", expiry_str)
    if m:
        month = int(m.group(1))
        year = int(m.group(2)) + 2000
        try:
            if month == 12:
                return datetime(year, 12, 31)
            else:
                return datetime(year, month + 1, 1) - timedelta(seconds=1)
        except Exception:
            pass

    # Try MM/YYYY or MM-YYYY
    m = re.match(r"^(\d{1,2})[/\-](\d{4})$", expiry_str)
    if m:
        month = int(m.group(1))
        year = int(m.group(2))
        try:
            if month == 12:
                return datetime(year, 12, 31)
            else:
                return datetime(year, month + 1, 1) - timedelta(seconds=1)
        except Exception:
            pass

    # Fallback format try
    for fmt in (
        "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%Y-%m-%dT%H:%M:%S",
        "%b-%y", "%b-%Y", "%b/%y", "%b/%Y", "%d-%b-%Y", "%d/%b/%Y", "%d-%b-%y",
        "%d/%b/%y", "%d %b %Y", "%d %b %y", "%b %d, %Y", "%B %Y", "%b %Y",
        "%Y-%m", "%Y/%m"
    ):
        try:
            return datetime.strptime(expiry_str, fmt)
        except ValueError:
            pass
            
    # Try dateutil parser as final fallback
    try:
        from dateutil.parser import parse as date_parse
        return date_parse(expiry_str)
    except Exception:
        pass
        
    return None

async def resolve_category_from_brand(brand: Optional[str], db) -> Optional[str]:
    if not brand:
        return None
    cleaned = brand.strip()
    if not cleaned:
        return None
    
    brand_regex = {"$regex": f"^{re.escape(cleaned)}$", "$options": "i"}
    existing = await db.categories.find_one({"name": brand_regex})
    if existing:
        return str(existing["_id"])
        
    res = await db.categories.insert_one({
        "name": cleaned,
        "description": f"Brand: {cleaned}",
        "is_active": True,
        "created_at": datetime.utcnow()
    })
    return str(res.inserted_id)


@router.get("/")
async def list_products(
    search: Optional[str] = Query(None),
    category_id: Optional[str] = Query(None),
    low_stock: Optional[bool] = Query(None),
    is_active: Optional[bool] = Query(True),
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=200),
    sort_by: Optional[str] = Query(None),
    sort_order: int = Query(1),
    db = Depends(get_database),
    current_user = Depends(require_permission("can_view_products"))
):
    query = {}
    if is_active is not None:
        query["is_active"] = is_active
    if category_id:
        query["category_id"] = category_id
    if search:
        import re
        escaped_search = re.escape(search)
        query["$or"] = [
            {"name": {"$regex": f"(^|\\s){escaped_search}", "$options": "i"}},
            {"sku": {"$regex": f"^{escaped_search}", "$options": "i"}},
            {"barcode": {"$regex": f"^{escaped_search}", "$options": "i"}},
            {"brand": {"$regex": f"(^|\\s){escaped_search}", "$options": "i"}}
        ]
    if low_stock:
        query["$expr"] = {"$lte": ["$current_stock", {"$ifNull": ["$min_stock_alert", 10.0]}]}

    sort_field = "name"
    if sort_by == "stock":
        sort_field = "current_stock"
    elif sort_by == "purchase":
        sort_field = "purchase_price"
    elif sort_by == "sale":
        sort_field = "selling_price"
    elif sort_by == "expiry":
        sort_field = "expiry"

    total = await db.products.count_documents(query)
    skip = (page - 1) * limit
    sort_list = [(sort_field, sort_order)]
    if sort_field != "name":
        sort_list.append(("name", 1))
    products = await db.products.find(query).skip(skip).limit(limit).sort(sort_list).to_list(limit)



    # Enrich with category names
    category_ids = list({p["category_id"] for p in products if p.get("category_id")})
    categories = {}
    if category_ids:
        cat_docs = await db.categories.find(
            {"_id": {"$in": [ObjectId(cid) for cid in category_ids if ObjectId.is_valid(cid)]}}
        ).to_list(1000)
        categories = {str(c["_id"]): c["name"] for c in cat_docs}

    result = []
    for p in products:
        doc = serialize_doc(p)
        doc["category_name"] = categories.get(doc.get("category_id", ""), "")
        result.append(doc)

    return {"items": result, "total": total, "page": page, "limit": limit}

@router.get("/search")
async def search_products(
    q: Optional[str] = Query(None),
    limit: int = Query(20, ge=1, le=100),
    db = Depends(get_database),
    current_user = Depends(require_permission("can_view_products"))
):
    """Fast product search for billing - returns minimal data."""
    import re
    projection = {
        "name": 1, "sku": 1, "barcode": 1, "selling_price": 1,
        "wholesale_price": 1, "mrp": 1, "gst_rate": 1, "unit": 1,
        "current_stock": 1, "hsn_code": 1, "purchase_price": 1,
        "brand": 1
    }

    if not q:
        # Return top active products if search query is empty
        products = await db.products.find(
            {"is_active": True},
            projection
        ).limit(limit).to_list(limit)
        return [serialize_doc(p) for p in products]

    escaped_q = re.escape(q)

    # Stage 1: Prefix-anchored search (lightning fast index scan)
    prefix_query = {
        "is_active": True,
        "$or": [
            {"name": {"$regex": f"^{escaped_q}", "$options": "i"}},
            {"sku": {"$regex": f"^{escaped_q}", "$options": "i"}},
            {"barcode": {"$regex": f"^{escaped_q}", "$options": "i"}},
        ]
    }
    products = await db.products.find(prefix_query, projection).limit(limit).to_list(limit)
    
    # Stage 2: Fallback to substring search (only if we got fewer results than limit)
    if len(products) < limit:
        remaining = limit - len(products)
        already_found_ids = {p["_id"] for p in products}
        
        substring_query = {
            "is_active": True,
            "_id": {"$nin": list(already_found_ids)},
            "$or": [
                {"name": {"$regex": f"(^|\\s){escaped_q}", "$options": "i"}},
                {"sku": {"$regex": f"^{escaped_q}", "$options": "i"}},
                {"barcode": {"$regex": f"^{escaped_q}", "$options": "i"}},
            ]
        }
        fallback_products = await db.products.find(substring_query, projection).limit(remaining).to_list(remaining)
        products.extend(fallback_products)

    return [serialize_doc(p) for p in products]

@router.post("/bulk-delete")
async def bulk_delete_products(
    payload: dict,
    db = Depends(get_database),
    current_user = Depends(require_permission("can_manage_products"))
):
    ids = payload.get("ids", [])
    object_ids = []
    for id_str in ids:
        try:
            object_ids.append(ObjectId(id_str))
        except:
            pass
            
    if not object_ids:
        raise HTTPException(status_code=400, detail="No valid IDs provided")
        
    await db.products.update_many(
        {"_id": {"$in": object_ids}},
        {"$set": {"is_active": False, "updated_at": datetime.utcnow()}}
    )
    return {"message": f"Successfully deleted {len(object_ids)} products"}

@router.get("/barcode/{barcode}")
async def get_by_barcode(
    barcode: str,
    db = Depends(get_database),
    current_user = Depends(require_permission("can_view_products"))
):
    product = await db.products.find_one({"barcode": barcode, "is_active": True})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    return serialize_doc(product)

@router.get("/{product_id}")
async def get_product(
    product_id: str,
    db = Depends(get_database),
    current_user = Depends(require_permission("can_view_products"))
):
    product = await db.products.find_one({"_id": ObjectId(product_id)})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    doc = serialize_doc(product)
    if doc.get("category_id"):
        cat = await db.categories.find_one({"_id": ObjectId(doc["category_id"])})
        doc["category_name"] = cat["name"] if cat else ""
    return doc

async def _create_product_internal(
    product_data: ProductCreate,
    db,
    current_user
):
    import re
    
    # 1. Clean fields: remove extra spaces
    cleaned_name = " ".join(product_data.name.split()).strip()
    cleaned_sku = " ".join(product_data.sku.split()).strip() if product_data.sku else None
    cleaned_barcode = " ".join(product_data.barcode.split()).strip() if product_data.barcode else None
    
    # 2. Check for duplicate/existing matching records
    # Check explicit matched_product_id link first
    existing = None
    if product_data.matched_product_id:
        try:
            existing = await db.products.find_one({"_id": ObjectId(product_data.matched_product_id)})
        except:
            pass
            
    # Match by Name (check active first, then inactive)
    if not existing and cleaned_name:
        name_regex = {"$regex": f"^{re.escape(cleaned_name)}$", "$options": "i"}
        existing = await db.products.find_one({"name": name_regex, "is_active": True})
        if not existing:
            existing = await db.products.find_one({"name": name_regex, "is_active": False})
            
    # Match by SKU (check active first, then inactive)
    if not existing and cleaned_sku:
        existing = await db.products.find_one({"sku": cleaned_sku, "is_active": True})
        if not existing:
            existing = await db.products.find_one({"sku": cleaned_sku, "is_active": False})
            
    # Match by Barcode (check active first, then inactive)
    if not existing and cleaned_barcode:
        existing = await db.products.find_one({"barcode": cleaned_barcode, "is_active": True})
        if not existing:
            existing = await db.products.find_one({"barcode": cleaned_barcode, "is_active": False})
            
    # Save correction to dataset automatically if user corrected it
    if product_data.raw_name and product_data.matched_product_id:
        raw_clean = product_data.raw_name.strip().lower()
        matched_prod = existing or await db.products.find_one({"_id": ObjectId(product_data.matched_product_id)})
        if matched_prod and raw_clean and matched_prod["name"].lower() != raw_clean:
            await db.corrections_dataset.update_one(
                {"raw_name": raw_clean},
                {
                    "$set": {
                        "raw_name": raw_clean,
                        "matched_product_id": str(matched_prod["_id"]),
                        "matched_product_name": matched_prod["name"],
                        "updated_at": datetime.utcnow()
                    },
                    "$setOnInsert": {
                        "created_at": datetime.utcnow()
                    }
                },
                upsert=True
            )
            
    if existing:
        # MERGE PRODUCT DETAILS AND STOCK
        old_stock = existing.get("current_stock", 0.0)
        add_stock = product_data.opening_stock
        new_stock = old_stock + add_stock
        
        now = datetime.utcnow()
        update_dict = {
            "name": cleaned_name,
            "current_stock": new_stock,
            "is_active": True,  # Reactivate in case it was inactive
            "updated_at": now
        }
        
        # Update latest prices/rates
        if product_data.purchase_price:
            update_dict["purchase_price"] = product_data.purchase_price
        if product_data.selling_price:
            update_dict["selling_price"] = product_data.selling_price
        if product_data.mrp:
            update_dict["mrp"] = product_data.mrp
        if product_data.wholesale_price:
            update_dict["wholesale_price"] = product_data.wholesale_price
        if product_data.hsn_code:
            update_dict["hsn_code"] = product_data.hsn_code.strip() if product_data.hsn_code else None
        if product_data.gst_rate is not None:
            update_dict["gst_rate"] = product_data.gst_rate
        if product_data.pack:
            update_dict["pack"] = product_data.pack.strip() if product_data.pack else None
            
        # Add up cases and final amount
        old_cases = existing.get("cases") or 0.0
        update_dict["cases"] = old_cases + (product_data.cases or 0.0)
        
        old_amount = existing.get("final_amount") or 0.0
        update_dict["final_amount"] = old_amount + (product_data.final_amount or 0.0)

        # Merge batch and expiry fields
        old_batch = existing.get("batch") or ""
        new_batch = product_data.batch or ""
        if new_batch and new_batch not in old_batch:
            update_dict["batch"] = f"{old_batch}, {new_batch}" if old_batch else new_batch

        old_expiry = existing.get("expiry") or ""
        new_expiry = product_data.expiry or ""
        if new_expiry and new_expiry not in old_expiry:
            update_dict["expiry"] = f"{old_expiry}, {new_expiry}" if old_expiry else new_expiry
        
        # Append description batch details
        old_desc = existing.get("description") or ""
        new_desc = product_data.description or ""
        if new_desc and new_desc not in old_desc:
            if old_desc:
                update_dict["description"] = f"{old_desc} | {new_desc}"
            else:
                update_dict["description"] = new_desc
                
        # Preserve barcode or SKU if existing had it, otherwise use new if provided
        if not existing.get("sku") and cleaned_sku:
            update_dict["sku"] = cleaned_sku
        if not existing.get("barcode") and cleaned_barcode:
            update_dict["barcode"] = cleaned_barcode
            
        # Resolve category from brand if brand is provided, otherwise keep specified category_id
        if product_data.brand and product_data.brand.strip():
            update_dict["brand"] = product_data.brand.strip()
            cat_id = await resolve_category_from_brand(product_data.brand, db)
            if cat_id:
                update_dict["category_id"] = cat_id
        elif product_data.category_id:
            update_dict["category_id"] = product_data.category_id
            
        await db.products.update_one({"_id": existing["_id"]}, {"$set": update_dict})
        
        # Log stock adjustment
        if add_stock > 0:
            await db.stock_logs.insert_one({
                "product_id": str(existing["_id"]),
                "product_name": existing["name"],
                "type": "adjustment",
                "quantity": add_stock,
                "before_stock": old_stock,
                "after_stock": new_stock,
                "reference": "Import Merge",
                "created_by": str(current_user["_id"]),
                "created_at": now
            })
            
            # Upsert into batches for tracking
            batch_no = product_data.batch.strip() if (product_data.batch and product_data.batch.strip()) else "DEFAULT"
            expiry_dt = parse_expiry_string(product_data.expiry)
            await db.batches.update_one(
                {"product_id": str(existing["_id"]), "batch_no": batch_no},
                {
                    "$inc": {"current_stock": add_stock},
                    "$set": {
                        "expiry": expiry_dt,
                        "purchase_price": product_data.purchase_price,
                        "updated_at": now
                    },
                    "$setOnInsert": {
                        "created_at": now
                    }
                },
                upsert=True
            )
            
        return {"message": "Product stock merged", "id": str(existing["_id"])}

    # 3. Double-check SKU uniqueness for new product creation (against other products)
    if cleaned_sku:
        existing_sku = await db.products.find_one({"sku": cleaned_sku})
        if existing_sku:
            raise HTTPException(status_code=400, detail="SKU already exists")

    # 4. Create new product
    now = datetime.utcnow()
    product_dict = product_data.dict()
    product_dict["name"] = cleaned_name
    product_dict["sku"] = cleaned_sku
    product_dict["barcode"] = cleaned_barcode
    if not product_dict.get("sku"):
        product_dict.pop("sku", None)
    if not product_dict.get("barcode"):
        product_dict.pop("barcode", None)
        
    # Resolve category from brand if brand is provided, otherwise keep specified category_id
    if product_data.brand and product_data.brand.strip():
        product_dict["brand"] = product_data.brand.strip()
        cat_id = await resolve_category_from_brand(product_data.brand, db)
        if cat_id:
            product_dict["category_id"] = cat_id
    elif product_data.category_id:
        product_dict["category_id"] = product_data.category_id
        
    product_dict["current_stock"] = product_data.opening_stock
    product_dict["created_at"] = now
    product_dict["updated_at"] = now
    product_dict["created_by"] = str(current_user["_id"])

    result = await db.products.insert_one(product_dict)

    # Log opening stock
    if product_data.opening_stock > 0:
        await db.stock_logs.insert_one({
            "product_id": str(result.inserted_id),
            "product_name": cleaned_name,
            "type": "opening",
            "quantity": product_data.opening_stock,
            "before_stock": 0,
            "after_stock": product_data.opening_stock,
            "reference": "Opening Stock",
            "created_by": str(current_user["_id"]),
            "created_at": now
        })
        
        # Upsert into batches for tracking
        batch_no = product_data.batch.strip() if (product_data.batch and product_data.batch.strip()) else "DEFAULT"
        expiry_dt = parse_expiry_string(product_data.expiry)
        await db.batches.update_one(
            {"product_id": str(result.inserted_id), "batch_no": batch_no},
            {
                "$inc": {"current_stock": product_data.opening_stock},
                "$set": {
                    "expiry": expiry_dt,
                    "purchase_price": product_data.purchase_price,
                    "updated_at": now
                },
                "$setOnInsert": {
                    "created_at": now
                }
            },
            upsert=True
        )

    return {"message": "Product created", "id": str(result.inserted_id)}

@router.post("/")
async def create_product(
    product_data: ProductCreate,
    db = Depends(get_database),
    current_user = Depends(require_permission("can_manage_products"))
):
    return await _create_product_internal(product_data, db, current_user)

@router.post("/bulk")
async def create_products_bulk(
    products_data: list[ProductCreate],
    db = Depends(get_database),
    current_user = Depends(require_permission("can_manage_products"))
):
    results = []
    errors = []
    for idx, item in enumerate(products_data):
        try:
            res = await _create_product_internal(item, db, current_user)
            results.append({"index": idx, "name": item.name, "message": res.get("message", "Product created"), "id": res["id"]})
        except HTTPException as he:
            errors.append({"index": idx, "name": item.name, "error": he.detail})
        except Exception as e:
            errors.append({"index": idx, "name": item.name, "error": str(e)})
            
    return {"results": results, "errors": errors}


@router.put("/{product_id}")
async def update_product(
    product_id: str,
    product_data: ProductUpdate,
    db = Depends(get_database),
    current_user = Depends(require_permission("can_manage_products"))
):
    update_dict = {k: v for k, v in product_data.dict().items() if v is not None}
    if not update_dict:
        raise HTTPException(status_code=400, detail="No data to update")

    # Resolve category from brand if brand is being updated
    if "brand" in update_dict:
        brand_val = update_dict["brand"]
        brand_str = brand_val.strip() if brand_val else ""
        if brand_str:
            update_dict["brand"] = brand_str
            cat_id = await resolve_category_from_brand(brand_str, db)
            if cat_id:
                update_dict["category_id"] = cat_id
        else:
            update_dict["brand"] = None

    update_dict["updated_at"] = datetime.utcnow()
    result = await db.products.update_one(
        {"_id": ObjectId(product_id)},
        {"$set": update_dict}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Product not found")
    return {"message": "Product updated"}

@router.delete("/{product_id}")
async def delete_product(
    product_id: str,
    db = Depends(get_database),
    current_user = Depends(require_permission("can_manage_products"))
):
    result = await db.products.update_one(
        {"_id": ObjectId(product_id)},
        {"$set": {"is_active": False, "updated_at": datetime.utcnow()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Product not found")
    return {"message": "Product deleted"}

@router.post("/bulk-import")
async def bulk_import_products(
    file: UploadFile = File(...),
    db = Depends(get_database),
    current_user = Depends(require_permission("can_manage_products"))
):
    """Import products from Excel file."""
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only Excel files supported")

    import openpyxl
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    headers = [str(cell.value).strip().lower().replace(" ", "_") for cell in ws[1]]
    imported = 0
    errors = []
    now = datetime.utcnow()

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            row_data = dict(zip(headers, row))
            if not row_data.get("name"):
                continue

            brand_val = row_data.get("brand")
            brand_str = str(brand_val).strip() if brand_val is not None else ""
            if brand_str.lower() in ("none", "nan", ""):
                brand_str = None
            else:
                brand_str = " ".join(brand_str.split())

            product_dict = {
                "name": str(row_data.get("name", "")),
                "sku": row_data.get("sku") or row_data.get("sku_code"),
                "barcode": row_data.get("barcode"),
                "brand": brand_str,
                "unit": str(row_data.get("unit", "PCS")),
                "hsn_code": row_data.get("hsn_code") or row_data.get("hsn"),
                "gst_rate": float(row_data.get("gst_rate", 18) or 18),
                "purchase_price": float(row_data.get("purchase_price", 0) or 0),
                "selling_price": float(row_data.get("selling_price", 0) or 0),
                "mrp": float(row_data.get("mrp", 0) or 0) or None,
                "wholesale_price": float(row_data.get("wholesale_price", 0) or 0) or None,
                "min_stock_alert": float(row_data.get("min_stock_alert", 10) or 10),
                "current_stock": float(row_data.get("opening_stock", 0) or 0),
                "opening_stock": float(row_data.get("opening_stock", 0) or 0),
                "is_active": True,
                "created_at": now,
                "updated_at": now,
                "created_by": str(current_user["_id"])
            }
            
            if brand_str:
                cat_id = await resolve_category_from_brand(brand_str, db)
                if cat_id:
                    product_dict["category_id"] = cat_id
            else:
                product_dict["category_id"] = None

            # Upsert by SKU or name
            filter_q = {}
            if product_dict["sku"]:
                filter_q["sku"] = product_dict["sku"]
            else:
                filter_q["name"] = product_dict["name"]

            await db.products.update_one(filter_q, {"$set": product_dict}, upsert=True)
            
            # Sync batch for Excel bulk import
            if product_dict["opening_stock"] > 0:
                prod_doc = await db.products.find_one(filter_q)
                if prod_doc:
                    # Log opening stock
                    await db.stock_logs.insert_one({
                        "product_id": str(prod_doc["_id"]),
                        "product_name": prod_doc["name"],
                        "type": "opening",
                        "quantity": product_dict["opening_stock"],
                        "before_stock": 0,
                        "after_stock": product_dict["opening_stock"],
                        "reference": "Excel Import Opening",
                        "created_by": str(current_user["_id"]),
                        "created_at": now
                    })
                    
                    batch_no = str(row_data.get("batch", "DEFAULT")).strip() or "DEFAULT"
                    expiry_str = str(row_data.get("expiry", "")).strip() or None
                    expiry_dt = parse_expiry_string(expiry_str)
                    
                    await db.batches.update_one(
                        {"product_id": str(prod_doc["_id"]), "batch_no": batch_no},
                        {
                            "$inc": {"current_stock": product_dict["opening_stock"]},
                            "$set": {
                                "expiry": expiry_dt,
                                "purchase_price": product_dict["purchase_price"],
                                "updated_at": now
                            },
                            "$setOnInsert": {
                                "created_at": now
                            }
                        },
                        upsert=True
                    )
            imported += 1
        except Exception as e:
            errors.append({"row": row_idx, "error": str(e)})

    return {"imported": imported, "errors": errors}

def _process_ocr_blocking(contents: bytes, filename_lower: str, content_type: str) -> list:
    import pytesseract
    from PIL import Image
    import io
    import re
    import os
    import pypdf
    import gc

    # Set tesseract path (including fallback for standard Linux/Debian path in Docker)
    tesseract_paths = [
        '/usr/bin/tesseract',
        '/opt/homebrew/bin/tesseract',
        '/usr/local/bin/tesseract',
        'tesseract'
    ]
    for path in tesseract_paths:
        if os.path.exists(path) or path == 'tesseract':
            pytesseract.pytesseract.tesseract_cmd = path
            break

    # Setup Pillow Resampling fallback compatibility for older PIL versions
    try:
        _ = Image.Resampling
    except AttributeError:
        class DummyResampling:
            LANCZOS = getattr(Image, 'LANCZOS', getattr(Image, 'ANTIALIAS', 1))
            BILINEAR = getattr(Image, 'BILINEAR', 2)
        Image.Resampling = DummyResampling

    is_pdf = False
    if filename_lower.endswith(".pdf") or content_type == "application/pdf" or contents.startswith(b"%PDF"):
        is_pdf = True
        
    # Save a debug copy to inspect layout
    try:
        debug_ext = ".pdf" if is_pdf else ".png"
        debug_path = f"static/invoices/debug_upload{debug_ext}"
        os.makedirs(os.path.dirname(debug_path), exist_ok=True)
        with open(debug_path, "wb") as f_debug:
            f_debug.write(contents)
    except Exception as e:
        print(f"Failed to save debug upload: {e}")

    if is_pdf:
        try:
            reader = pypdf.PdfReader(io.BytesIO(contents))
            text = ""
            for page in reader.pages:
                page_text = ""
                try:
                    page_text = page.extract_text(extraction_mode="layout") or ""
                except Exception:
                    pass
                # Fallback if layout mode fails or returns empty/short text
                if len(page_text.strip()) < 100:
                    page_text = page.extract_text() or ""
                text += page_text
            
            # If direct text extraction is empty/short, it's a scanned PDF:
            # Extract embedded images and run local OCR
            if len(text.strip()) < 100:
                ocr_text_parts = []
                for page in reader.pages[:3]:
                    for img_obj in page.images:
                        try:
                            img_data = img_obj.data
                            img = Image.open(io.BytesIO(img_data))
                            
                            # Only OCR images large enough to be actual page scans (ignores logos/icons to save RAM/time)
                            if img.width < 400 or img.height < 400:
                                img.close()
                                continue
                            
                            # Resize image to exactly 1000 width to optimize OCR readability and keep memory low
                            if img.width != 1000:
                                ratio = 1000.0 / img.width
                                new_height = int(img.height * ratio)
                                resample_filter = Image.Resampling.LANCZOS if img.width < 1000 else Image.Resampling.BILINEAR
                                old_img = img
                                img = old_img.resize((1000, new_height), resample_filter)
                                old_img.close()
                            
                            # Preprocess image (Grayscale + Enhance Contrast to save memory and improve OCR)
                            img_gray = img.convert('L')
                            img.close()
                            
                            from PIL import ImageEnhance
                            enhancer = ImageEnhance.Contrast(img_gray)
                            img_enhanced = enhancer.enhance(2.0)
                            
                            ocr_text_parts.append(pytesseract.image_to_string(img_enhanced, config='--psm 6'))
                            
                            img_enhanced.close()
                            img_gray.close()
                            del img_gray
                            del img_data
                            gc.collect()
                        except Exception as ocr_err:
                            print(f"Failed to OCR PDF image object: {ocr_err}")
                if ocr_text_parts:
                    text = "\n".join(ocr_text_parts)
        except Exception as e:
            raise ValueError(f"Failed to parse PDF file: {str(e)}")
    else:
        try:
            img = Image.open(io.BytesIO(contents))
            
            # Resize image to exactly 1000 width to optimize OCR readability and keep memory low
            if img.width != 1000:
                ratio = 1000.0 / img.width
                new_height = int(img.height * ratio)
                resample_filter = Image.Resampling.LANCZOS if img.width < 1000 else Image.Resampling.BILINEAR
                old_img = img
                img = old_img.resize((1000, new_height), resample_filter)
                old_img.close()
            
            # Preprocess the image (Grayscale + Enhance Contrast to save memory and improve OCR)
            img_gray = img.convert('L')
            img.close()
            
            from PIL import ImageEnhance
            enhancer = ImageEnhance.Contrast(img_gray)
            img_enhanced = enhancer.enhance(2.0)
            
            # Run OCR
            text = pytesseract.image_to_string(img_enhanced, config='--psm 6')
            
            img_enhanced.close()
            img_gray.close()
            del img_gray
            gc.collect()
        except Exception as e:
            raise ValueError(f"Invalid image file or OCR failed: {str(e)}")
    
    # Parse extracted text using robust pattern-based algorithm
    def is_batch_token(token):
        # Tokens with decimal points are prices/amounts, never batch numbers
        if '.' in token:
            return False
        if token.isdigit():
            return True
        if re.search(r'\d', token) and re.search(r'[A-Za-z]', token):
            if any(x in token.upper() for x in ['ML', 'UNIT', 'PCS', 'GM', 'KG', 'TAB', 'CAP', 'MM', 'INCH', 'CM', 'BOX', 'PACK']):
                return False
            # Cannula size suffix (e.g. VENO-20, VENO-18)
            if re.search(r'-\d{1,2}$', token):
                return False
            return True
        if re.match(r'^\d{4}\b', token):
            return True
        return False

    products = []
    lines = text.splitlines()
    
    # Keep track of expected SN (approximate counter for Yash Surgical)
    expected_sn = 1
    
    for line in lines:
        # Replace common separator characters with spaces (keep parentheses)
        cleaned = line.replace('|', ' ').replace('[', ' ').replace(']', ' ')
        cleaned = cleaned.replace('{', ' ').replace('}', ' ')
        cleaned = cleaned.replace(',', ' ').replace(';', ' ')
        
        tokens = cleaned.split()
        if len(tokens) < 6:
            continue
            
        # 1. Find HSN index (rightmost token of length 6 to 10)
        hsn_index = -1
        for idx in range(len(tokens) - 1, -1, -1):
            token = tokens[idx]
            if '.' in token:
                continue
            clean_tok = re.sub(r'[^\w\$\#]', '', token)
            if len(clean_tok) >= 6 and len(clean_tok) <= 10:
                if re.search(r'\d', clean_tok):
                    hsn_index = idx
                    break
                    
        if hsn_index == -1:
            # Fallback to search for known HSN prefixes
            for idx in range(len(tokens) - 1, -1, -1):
                token = tokens[idx]
                if any(x in token for x in ['3004', '300', '9018', '4015', '1902']):
                    hsn_index = idx
                    break
                    
        if hsn_index == -1 or hsn_index < 2:
            continue
            
        hsn = re.sub(r'[^\w]', '', tokens[hsn_index]) # Clean HSN
        
        # 2. Detect Format (A vs B)
        # Find Expiry date to the left of HSN
        exp_index = -1
        for idx in range(hsn_index - 1, -1, -1):
            tok = tokens[idx]
            if '/' in tok or (len(tok) == 4 and tok.isdigit() and int(tok[:2]) <= 12 and int(tok[2:]) >= 24 and int(tok[2:]) <= 40):
                exp_index = idx
                break
                
        if exp_index == -1:
            # Fallback to look for a token that has 4 digits or looks like exp
            for idx in range(hsn_index - 1, -1, -1):
                tok = tokens[idx]
                if len(tok) == 4 and tok.isdigit() and int(tok[:2]) <= 12 and int(tok[2:]) >= 24 and int(tok[2:]) <= 40:
                    exp_index = idx
                    break
                    
        if exp_index == -1:
            exp_index = hsn_index - 1 # Fallback
            
        # If expiry index is close to HSN index (distance <= 2), it is Format B
        is_format_b = (hsn_index - exp_index) <= 2
        
        try:
            if is_format_b:
                # Format B (Yash Surgical style)
                # Find MRP and Rate to the right of HSN
                right_tokens = tokens[hsn_index + 1:]
                decimal_right = []
                for t in right_tokens:
                    if ':' in t or '/' in t or re.search(r'[A-Za-z]', t):
                        continue
                    t_clean = re.sub(r'[^\d\.]', '', t)
                    if t_clean and ('.' in t_clean or t_clean.isdigit()):
                        try:
                            decimal_right.append(float(t_clean))
                        except:
                            pass
                mrp = decimal_right[0] if len(decimal_right) > 0 else 0.0
                rate = decimal_right[1] if len(decimal_right) > 1 else 0.0
                amount = decimal_right[5] if len(decimal_right) > 5 else (decimal_right[4] if len(decimal_right) > 4 else 0.0)
                
                exp = tokens[exp_index]
                exp_index_for_batch = exp_index
                if exp.startswith('/') and exp_index - 1 >= 0:
                    prev_tok = tokens[exp_index - 1]
                    if len(prev_tok) <= 2 or (prev_tok.isdigit() and int(prev_tok) <= 12):
                        exp = prev_tok + exp
                        exp_index_for_batch = exp_index - 1
                
                # Find Batch tokens (up to 2 tokens immediately left of exp_index_for_batch)
                batch_tokens = []
                for idx in range(exp_index_for_batch - 1, max(1, exp_index_for_batch - 3), -1):
                    tok = tokens[idx]
                    if is_batch_token(tok):
                        batch_tokens.insert(0, tok)
                    else:
                        break
                        
                batch = " ".join(batch_tokens) if batch_tokens else ""
                
                # Product Name is everything between the pack/quantity and the batch tokens
                first_tok = tokens[0]
                qty_val = 1.0
                pack_val = tokens[2]
                name_start_idx = 3
                
                # Detect merged SN + Qty in token 0
                if re.search(r'\d+[\.\$]?\d+\.\d{2}$', first_tok) or (expected_sn and first_tok.startswith(str(expected_sn)) and len(first_tok) > len(str(expected_sn)) + 2):
                    # Merged!
                    if rate > 0 and amount > 0:
                        qty_val = round(amount / rate)
                    else:
                        match = re.search(r'\d+\.\d{2}$', first_tok)
                        if match:
                            qty_val = float(match.group())
                    
                    pack_val = tokens[1]
                    name_start_idx = 2
                    expected_sn += 1
                else:
                    qty_str = tokens[1]
                    qty_str_clean = re.sub(r'[^\d\.]', '', qty_str)
                    qty_val = float(qty_str_clean) if qty_str_clean else 1.0
                    
                    # Verify using math if possible
                    if rate > 0 and amount > 0:
                        calc_qty = round(amount / rate)
                        if abs(calc_qty - qty_val) > 2:
                            qty_val = calc_qty
                            
                    # Update expected SN if token 0 matches integer
                    clean_sn = re.sub(r'[^\d]', '', first_tok)
                    if clean_sn.isdigit():
                        expected_sn = int(clean_sn) + 1
                
                # Name tokens
                batch_start_idx = exp_index_for_batch - len(batch_tokens)
                name_tokens = tokens[name_start_idx:batch_start_idx]
                
                if len(name_tokens) > 0 and name_tokens[0] == pack_val:
                    name_tokens = name_tokens[1:]
                name = " ".join(name_tokens).strip()
                
            else:
                # Format A (R B Healthcare style)
                exp = tokens[exp_index]
                exp_index_for_batch = exp_index
                if exp.startswith('/') and exp_index - 1 >= 0:
                    prev_tok = tokens[exp_index - 1]
                    if len(prev_tok) <= 2 or (prev_tok.isdigit() and int(prev_tok) <= 12):
                        exp = prev_tok + exp
                        exp_index_for_batch = exp_index - 1

                # Batch is to the left of exp_index_for_batch
                batch = tokens[exp_index_for_batch - 1]
                
                # Price and Qty columns are between exp_index and hsn_index
                mid_tokens = tokens[exp_index + 1:hsn_index]
                qty_val = 1.0
                mrp = 0.0
                rate = 0.0
                
                if len(mid_tokens) >= 3:
                    qty_val = float(re.sub(r'[^\d\.]', '', mid_tokens[0]))
                    mrp = float(re.sub(r'[^\d\.]', '', mid_tokens[1]))
                    rate = float(re.sub(r'[^\d\.-]', '', mid_tokens[2]).replace('-', '.'))
                elif len(mid_tokens) == 2:
                    mrp = float(re.sub(r'[^\d\.]', '', mid_tokens[0]))
                    rate = float(re.sub(r'[^\d\.-]', '', mid_tokens[1]).replace('-', '.'))
                
                # Product Name is everything before batch
                name_tokens = tokens[:exp_index_for_batch - 1]
                if len(name_tokens) > 1 and (name_tokens[0].isdigit() or len(name_tokens[0]) == 1 or name_tokens[0].endswith('.')):
                    name_tokens = name_tokens[1:]
                name = " ".join(name_tokens).strip()
                pack_val = None

            # Clean up product name typos
            name = re.sub(r'\bSOOML\b', '500ML', name, flags=re.IGNORECASE)
            name = re.sub(r'\bSOOM\b', '500ML', name, flags=re.IGNORECASE)
            name = re.sub(r'\bSOO\b', '500ML', name, flags=re.IGNORECASE)
            name = re.sub(r'\b100M\b', '100ML', name, flags=re.IGNORECASE)
            name = name.replace('ELEP', 'ELE(P)')
            name = re.sub(r'\bCP\.', 'CP', name)
            
            # Clean common packing indicators from start/end of name
            name_patterns = [
                r'^\s*(?:\d+\s*[xX*\-]?\s*)?(?:UNIT|PCS|BOX|BAG)S?\b',
                r'\b(?:\d+\s*[xX*\-]?\s*)?(?:UNIT|PCS|BOX|BAG)S?\s*$',
                r'^\s*\d*[xX*]\d+[a-zA-Z]?\b',
                r'\b\d*[xX*]\d+[a-zA-Z]?\s*$',
                r'^\s*\d+\s*[\*xX\-/°]\s*\d+\b',
                r'\b\d+\s*[\*xX\-/°]\s*\d+\s*$',
            ]
            for pattern in name_patterns:
                name = re.sub(pattern, '', name, flags=re.IGNORECASE).strip()
            
            name = re.sub(r'\s+', ' ', name).strip()
            
            # Strip trailing pack multipliers safely
            name = re.sub(r'\b1\s*[\*xX/°\-o]\s*\d+\b\s*$', '', name, flags=re.IGNORECASE).strip()
            
            # Determine default pack size
            default_pack_size = 24
            if pack_val:
                nums = re.findall(r'\d+', pack_val)
                if len(nums) > 1:
                    default_pack_size = int(nums[-1])
                elif len(nums) == 1:
                    default_pack_size = int(nums[0])
            else:
                if '100ML' in name:
                    default_pack_size = 100
                elif '500ML' in name:
                    default_pack_size = 24
                    
            if not pack_val:
                pack_val = f"1*{default_pack_size}"
            
            # Clean up exp (e.g. 2728 -> 2/28, 4728 -> 4/28)
            if '/' not in exp and len(exp) >= 3:
                if re.match(r'^\d7\d{2}$', exp):
                    exp = f"{exp[0]}/{exp[2:]}"
                elif re.match(r'^\d{2}7\d{2}$', exp):
                    exp = f"{exp[:2]}/{exp[3:]}"
                elif exp.isdigit():
                    if len(exp) == 3:
                        exp = f"{exp[0]}/{exp[1:]}"
                    elif len(exp) == 4:
                        exp = f"{exp[:-2]}/{exp[-2:]}"
            
            # GST rate (usually at fixed offset, with fallback)
            gst = 5.0
            preferred_offsets = [4, 2] if is_format_b else [2, 4]
            for offset in preferred_offsets + [1, 3]:
                if hsn_index + offset < len(tokens):
                    try:
                        gst_str = re.sub(r'[^\d\.]', '', tokens[hsn_index + offset])
                        val = float(gst_str)
                        if val in [0.0, 3.0, 5.0, 12.0, 18.0, 28.0]:
                            gst = val
                            break
                        elif val <= 100.0 and gst == 5.0:
                            gst = val
                    except:
                        pass
            # Parse final amount from the end of the tokens list (before Cases)
            parsed_amount = None
            for offset in [-2, -1, -3]:
                if len(tokens) + offset >= 0:
                    tok = tokens[offset]
                    if '.' in tok:
                        try:
                            clean_tok = re.sub(r'[^\d\.]', '', tok)
                            val = float(clean_tok)
                            if val > 100.0 or parsed_amount is None:
                                parsed_amount = val
                        except:
                            pass
            
            # If we got a valid parsed amount, run mathematical self-correction
            if rate > 0 and parsed_amount is not None:
                try:
                    calculated_qty = round(parsed_amount / rate, 2)
                    if qty_val == 1.0 or abs(qty_val - calculated_qty) > 5.0:
                        qty_val = calculated_qty
                except:
                    pass

            # Dynamically calculate final amount
            final_amount = round(qty_val * rate, 2)
            
            # Cases column
            cases = None
            if not is_format_b:
                if len(tokens) >= 12:
                    try:
                        last_token = tokens[-1]
                        if '.' in last_token or float(last_token) > 200:
                            cases = round(qty_val / default_pack_size)
                        else:
                            cases = int(float(last_token))
                    except:
                        cases = round(qty_val / default_pack_size)
                else:
                    cases = round(qty_val / default_pack_size)
            
            if not name or len(name.strip()) < 2:
                continue
            if rate <= 0 and mrp <= 0:
                continue

            # Clean up month misreads (e.g. P/31 -> 2/31, p/31 -> 2/31, l/31 -> 1/31)
            if '/' in exp:
                parts = exp.split('/')
                month_part = parts[0].strip()
                year_part = parts[1].strip() if len(parts) > 1 else ""
                
                month_part = re.sub(r'^[pP]$', '2', month_part)
                month_part = re.sub(r'^[lIi|]$', '1', month_part)
                month_part = re.sub(r'^[sS]$', '5', month_part)
                
                if year_part:
                    exp = f"{month_part}/{year_part}"
                else:
                    exp = month_part
                
            item = {
                "name": name,
                "sku": None,
                "barcode": None,
                "brand": "YASH SURGICAL HOUSE" if is_format_b else "R B HEALTHCARE",
                "unit": "PCS",
                "hsn_code": hsn,
                "gst_rate": gst,
                "purchase_price": rate,
                "selling_price": mrp,
                "mrp": mrp,
                "wholesale_price": round(rate * 1.1, 2),
                "opening_stock": qty_val,
                "min_stock_alert": 10.0,
                "description": f"Batch: {batch}, Exp: {exp}",
                "is_active": True,
                "pack": pack_val,
                "cases": cases,
                "final_amount": final_amount,
                "batch": batch,
                "expiry": exp
            }
            products.append(item)
        except Exception:
            pass
            
    if not products:
        raise ValueError(
            "No products could be extracted. Please ensure the invoice is clear, "
            "well-lit, and matches the supported formats (Yash Surgical / RB Healthcare)."
        )
    gc.collect()
    return products

async def _analyze_invoice_with_gemini(contents: bytes, content_type: str) -> list:
    if not settings.GEMINI_API_KEY:
        raise ValueError("GEMINI_API_KEY is not set")
    
    # Map common extensions to standard mime types if content_type is generic/missing
    if not content_type or content_type == "application/octet-stream":
        content_type = "image/png"  # Default fallback
        
    encoded_image = base64.b64encode(contents).decode("utf-8")
    
    prompt = """You are an expert invoice processing assistant for a wholesale ERP.
Analyze the provided invoice image or PDF and extract the list of products/items being purchased.
For each item, extract and compute the following fields:
- name: The product name (e.g. "Veno-20 Cannula", "500ML Normal Saline"). Clean up obvious scan/OCR errors and typos, standardizing formatting (e.g. "SOOML" -> "500ML").
- sku: Product SKU if visible, otherwise null.
- barcode: Product barcode if visible, otherwise null.
- brand: The distributor/seller name at the top of the invoice (e.g. "YASH SURGICAL HOUSE", "R B HEALTHCARE", or others).
- unit: The unit of measurement (usually "PCS").
- hsn_code: The HSN code for the item.
- gst_rate: The GST percentage rate applied (e.g. 5.0, 12.0, 18.0, 28.0).
- purchase_price: The rate or purchase price per unit.
- selling_price: The MRP or selling price.
- mrp: The Maximum Retail Price (MRP).
- wholesale_price: The wholesale price, calculated as purchase_price * 1.1, rounded to 2 decimal places.
- opening_stock: The quantity purchased in this invoice.
- min_stock_alert: 10.0 (default value).
- pack: The packing size (e.g. "1*24", "1*100").
- cases: The number of cases (usually opening_stock divided by the packing size multiplier, e.g. if opening_stock is 240 and pack is 1*24, cases is 10).
- final_amount: The total purchase price for this item (opening_stock * purchase_price). Ensure the math is corrected if decimal points are missing or misread in the raw text (e.g., if Rate is 15.20 and Qty is 100, the amount is 1520.00).
- batch: The batch number of the item.
- expiry: The expiry date in format MM/YY or MM/YYYY (e.g. "05/28"). If raw date is like "P/31" or "0528", correct it to a valid date format.
- description: "Batch: {batch}, Exp: {expiry}"
- is_active: true (default value).

Verify all calculations:
1. Check that opening_stock * purchase_price matches final_amount within a reasonable margin. If there's a discrepancy, correct any misparsed decimal points (e.g., if rate is parsed as 1520 but amount is 1520.00 and quantity is 100, then rate should be corrected to 15.20).
2. If packing size is e.g. "1*24", make sure opening_stock is a multiple of 24 (or close to it) and matches the cases column correctly.
"""

    payload = {
        "contents": [
            {
                "parts": [
                    {"text": prompt},
                    {
                        "inlineData": {
                            "mimeType": content_type,
                            "data": encoded_image
                        }
                    }
                ]
            }
        ],
        "generationConfig": {
            "responseMimeType": "application/json",
            "responseSchema": {
                "type": "ARRAY",
                "items": {
                    "type": "OBJECT",
                    "properties": {
                        "name": {"type": "STRING"},
                        "sku": {"type": "STRING", "nullable": True},
                        "barcode": {"type": "STRING", "nullable": True},
                        "brand": {"type": "STRING"},
                        "unit": {"type": "STRING"},
                        "hsn_code": {"type": "STRING"},
                        "gst_rate": {"type": "NUMBER"},
                        "purchase_price": {"type": "NUMBER"},
                        "selling_price": {"type": "NUMBER"},
                        "mrp": {"type": "NUMBER"},
                        "wholesale_price": {"type": "NUMBER"},
                        "opening_stock": {"type": "NUMBER"},
                        "min_stock_alert": {"type": "NUMBER"},
                        "pack": {"type": "STRING"},
                        "cases": {"type": "NUMBER", "nullable": True},
                        "final_amount": {"type": "NUMBER"},
                        "batch": {"type": "STRING"},
                        "expiry": {"type": "STRING"},
                        "description": {"type": "STRING"},
                        "is_active": {"type": "BOOLEAN"}
                    },
                    "required": [
                        "name", "brand", "unit", "hsn_code", "gst_rate", 
                        "purchase_price", "selling_price", "mrp", 
                        "wholesale_price", "opening_stock", "min_stock_alert", 
                        "pack", "final_amount", "batch", "expiry", 
                        "description", "is_active"
                    ]
                }
            }
        }
    }

    url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-1.5-flash:generateContent?key={settings.GEMINI_API_KEY}"
    
    async with httpx.AsyncClient(timeout=30.0) as client:
        response = await client.post(url, json=payload)
        response.raise_for_status()
        resp_json = response.json()
        
        try:
            candidates = resp_json.get("candidates", [])
            if not candidates:
                raise ValueError("No candidates returned from Gemini API")
            text_response = candidates[0]["content"]["parts"][0]["text"]
            products = json.loads(text_response)
            
            validated_products = []
            for item in products:
                batch = item.get("batch") or ""
                expiry = item.get("expiry") or ""
                if not item.get("description"):
                    item["description"] = f"Batch: {batch}, Exp: {expiry}"
                
                item["sku"] = item.get("sku") or None
                item["barcode"] = item.get("barcode") or None
                item["cases"] = item.get("cases") or None
                
                validated_products.append(item)
                
            return validated_products
        except (KeyError, IndexError, json.JSONDecodeError) as e:
            raise ValueError(f"Failed to parse Gemini API response: {str(e)}")

def _validate_extracted_products(products: list) -> list:
    for item in products:
        errors = []
        
        # Math check: Qty * Rate = Amount
        qty = item.get("opening_stock")
        rate = item.get("purchase_price")
        amt = item.get("final_amount")
        
        if qty is not None and rate is not None:
            try:
                expected_amt = round(float(qty) * float(rate), 2)
                if amt is not None and abs(expected_amt - float(amt)) > 0.05:
                    errors.append(f"Qty * Rate ({expected_amt}) does not match Final Amount ({amt})")
            except:
                pass
                
        # GST rate verification
        gst = item.get("gst_rate")
        if gst is not None:
            try:
                gst_val = float(gst)
                if gst_val not in [0.0, 3.0, 5.0, 12.0, 18.0, 28.0]:
                    errors.append(f"Non-standard GST rate: {gst}%")
            except:
                errors.append("Invalid GST rate value")
                
        # Expiry format validation (MM/YY or MM/YYYY)
        exp = item.get("expiry")
        if exp:
            exp_str = str(exp).strip()
            if not re.match(r'^\d{1,2}[/\-]\d{2,4}$', exp_str):
                errors.append(f"Invalid expiry format: '{exp_str}' (Expected MM/YY or MM/YYYY)")
                
        # Batch verification
        batch = item.get("batch")
        if not batch or str(batch).strip() == "":
            errors.append("Batch number is missing")
            
        # Match confidence verification
        conf = item.get("confidence")
        if conf is not None:
            try:
                conf_val = float(conf)
                if conf_val < 0.6:
                    errors.append(f"Low matching confidence ({round(conf_val*100)}%) - Review mapped product")
            except:
                pass
        else:
            # Tesseract or fallback which has no confidence score
            if not item.get("matched_product_id"):
                errors.append("Unmapped product - Review suggested name")
            
        item["validation_errors"] = errors
    return products

async def _run_ocr_background(
    task_id: str,
    contents: bytes,
    filename_lower: str,
    content_type: str,
    db
):
    async with ocr_lock:
        try:
            products = None
            ai_used = False
            
            # Fetch active products and corrections for stateless AI matching
            db_products = []
            try:
                db_products_cursor = db.products.find({"is_active": True}, {"_id": 1, "name": 1})
                db_products = [{"id": str(p["_id"]), "name": p["name"]} for p in await db_products_cursor.to_list(10000)]
            except Exception as dbe:
                logger.error(f"Failed to fetch db_products for AI matching: {dbe}")
                
            corrections = []
            try:
                corrections_cursor = db.corrections_dataset.find({})
                corrections = [{"raw_name": c["raw_name"], "matched_product_id": c["matched_product_id"]} for c in await corrections_cursor.to_list(1000)]
            except Exception as ce:
                logger.error(f"Failed to fetch corrections for AI matching: {ce}")

            # Try self-hosted AI service if configured
            if settings.AI_SERVICE_URL:
                try:
                    logger.info("Attempting invoice analysis using self-hosted AI Service...")
                    files = {"file": (filename_lower, contents, content_type or "image/png")}
                    data = {
                        "db_products_str": json.dumps(db_products),
                        "corrections_str": json.dumps(corrections)
                    }
                    async with httpx.AsyncClient(timeout=45.0) as client:
                        resp = await client.post(f"{settings.AI_SERVICE_URL}/analyze-invoice", files=files, data=data)
                        resp.raise_for_status()
                        result_data = resp.json()
                        products = result_data.get("products", [])
                        ai_used = True
                        logger.info("Self-hosted AI analysis succeeded!")
                except Exception as ae:
                    logger.warning(f"Self-hosted AI service failed: {ae}. Falling back to local OCR...")
            
            # Local Tesseract Fallback
            if not ai_used or products is None:
                # Offload heavy local OCR/PDF parsing to thread pool
                products = await run_in_threadpool(
                    _process_ocr_blocking,
                    contents,
                    filename_lower,
                    content_type
                )
                
                # Perform simple local matching for fallback results
                for item in products:
                    item_name_lower = item["name"].strip().lower()
                    
                    # Direct check against corrections
                    matched_id = next((c["matched_product_id"] for c in corrections if c["raw_name"] == item_name_lower), None)
                    if matched_id:
                        matched_prod = next((p for p in db_products if p["id"] == matched_id), None)
                        if matched_prod:
                            item["matched_product_id"] = matched_prod["id"]
                            item["matched_product_name"] = matched_prod["name"]
                            item["confidence"] = 1.0
                            continue
                            
                    # Local fallback case-insensitive exact or prefix match
                    local_match = next((p for p in db_products if p["name"].lower() == item_name_lower or p["name"].lower().startswith(item_name_lower)), None)
                    if local_match:
                        item["matched_product_id"] = local_match["id"]
                        item["matched_product_name"] = local_match["name"]
                        item["confidence"] = 0.8
                    else:
                        item["matched_product_id"] = None
                        item["confidence"] = 0.0

            # Run validation checks on parsed products
            products = _validate_extracted_products(products)
                
            await db.ocr_tasks.update_one(
                {"_id": ObjectId(task_id)},
                {
                    "$set": {
                        "status": "completed",
                        "result": products,
                        "completed_at": datetime.utcnow()
                    }
                }
            )
        except ValueError as ve:
            await db.ocr_tasks.update_one(
                {"_id": ObjectId(task_id)},
                {
                    "$set": {
                        "status": "failed",
                        "error": str(ve),
                        "completed_at": datetime.utcnow()
                    }
                }
            )
        except Exception as e:
            logger.exception("Failed to analyze invoice in background")
            await db.ocr_tasks.update_one(
                {"_id": ObjectId(task_id)},
                {
                    "$set": {
                        "status": "failed",
                        "error": f"Internal error: {str(e)}",
                        "completed_at": datetime.utcnow()
                    }
                }
            )

@router.post("/import-image")
async def import_product_image(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    db = Depends(get_database),
    current_user = Depends(require_permission("can_manage_products"))
):
    try:
        contents = await file.read()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read file: {str(e)}")

    filename_lower = file.filename.lower()
    content_type = file.content_type or ""
    
    # Generate task ID
    task_id = str(ObjectId())
    
    # Create task record in MongoDB
    task_doc = {
        "_id": ObjectId(task_id),
        "status": "processing",
        "created_at": datetime.utcnow()
    }
    await db.ocr_tasks.insert_one(task_doc)
    
    # Register background task
    background_tasks.add_task(
        _run_ocr_background,
        task_id,
        contents,
        filename_lower,
        content_type,
        db
    )
    
    return {"task_id": task_id, "status": "processing"}

@router.get("/import-image/task/{task_id}")
async def get_import_task_status(
    task_id: str,
    db = Depends(get_database),
    current_user = Depends(require_permission("can_manage_products"))
):
    if not ObjectId.is_valid(task_id):
        raise HTTPException(status_code=400, detail="Invalid task ID")
        
    task = await db.ocr_tasks.find_one({"_id": ObjectId(task_id)})
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
        
    return {
        "status": task.get("status"),
        "result": task.get("result"),
        "error": task.get("error")
    }

@router.post("/corrections")
async def save_correction(
    payload: dict,
    db = Depends(get_database),
    current_user = Depends(require_permission("can_manage_products"))
):
    raw_name = payload.get("raw_name")
    matched_product_id = payload.get("matched_product_id")
    matched_product_name = payload.get("matched_product_name")
    
    if not raw_name or not matched_product_id:
        raise HTTPException(status_code=400, detail="raw_name and matched_product_id are required")
        
    now = datetime.utcnow()
    await db.corrections_dataset.update_one(
        {"raw_name": raw_name.strip().lower()},
        {
            "$set": {
                "raw_name": raw_name.strip().lower(),
                "matched_product_id": matched_product_id,
                "matched_product_name": matched_product_name,
                "updated_at": now
            },
            "$setOnInsert": {
                "created_at": now
            }
        },
        upsert=True
    )
    return {"status": "success", "message": "Correction saved"}

@router.post("/retrain-matching")
async def trigger_retrain(
    db = Depends(get_database),
    current_user = Depends(require_permission("can_manage_products"))
):
    corrections_cursor = db.corrections_dataset.find({})
    corrections = await corrections_cursor.to_list(1000)
    
    if len(corrections) < 5:
        raise HTTPException(status_code=400, detail="Need at least 5 corrections to run training")
        
    dataset = []
    for c in corrections:
        dataset.append({
            "anchor": c["raw_name"],
            "positive": c["matched_product_name"]
        })
        
    try:
        async with httpx.AsyncClient(timeout=60.0) as client:
            resp = await client.post(f"{settings.AI_SERVICE_URL}/retrain", json={"dataset": dataset})
            resp.raise_for_status()
            return resp.json()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to trigger AI retraining: {str(e)}")


